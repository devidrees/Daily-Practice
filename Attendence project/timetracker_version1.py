import tkinter as tk
from tkinter import ttk
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta

# Data storage in memory
time_records = []

# Function to calculate duration
def calculate_duration(time_in, time_out):
    duration = time_out - time_in
    hours, remainder = divmod(duration.seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{hours:02}:{minutes:02}:{seconds:02}"

# Function to handle IN button click
def time_in():
    global start_time, timer_running
    if timer_running:
        return  # Prevent multiple IN clicks
    start_time = datetime.now()
    timer_running = True
    time_records.append({"Date": start_time.date(), "Time In": start_time.time(), "Time Out": None, "Time Spent": None})
    update_timer()
    in_button.config(style="Green.TButton", state=tk.DISABLED)
    out_button.config(style="Light.TButton", state=tk.NORMAL)
    status_label.config(text=f"Time In recorded at {start_time.strftime('%H:%M:%S')}")

# Function to handle OUT button click
def time_out():
    global timer_running
    if not timer_running:
        status_label.config(text="You must clock IN first!")
        return
    end_time = datetime.now()
    duration = calculate_duration(start_time, end_time)
    time_records[-1]["Time Out"] = end_time.time()
    time_records[-1]["Time Spent"] = duration
    timer_running = False
    last_time_out_label.config(text=f"Last Time Out: {end_time.strftime('%H:%M:%S')}")
    last_spent_time_label.config(text=f"Last Spent Time: {duration}")
    out_button.config(style="Red.TButton", state=tk.DISABLED)
    in_button.config(style="Light.TButton", state=tk.NORMAL)
    status_label.config(text=f"Time Out recorded at {end_time.strftime('%H:%M:%S')}")
    update_history()

# Function to update the live IN timer
def update_timer():
    if timer_running:
        elapsed_time = datetime.now() - start_time
        hours, remainder = divmod(elapsed_time.seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        live_timer_label.config(text=f"Live IN Timer: {hours:02}:{minutes:02}:{seconds:02}")
        root.after(1000, update_timer)  # Update every second

# Function to save data to Excel
def save_data():
    if not time_records or not time_records[-1]["Time Out"]:
        status_label.config(text="You must clock OUT first!")
        return
    try:
        wb = load_workbook('time_tracker.xlsx')
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(['Date', 'Time In', 'Time Out', 'Time Spent'])
    
    # Insert new row before the last total row
    last_row = ws.max_row
    if ws.cell(row=last_row, column=1).value == "Total Time:":
        last_row -= 1
    
    for record in time_records:
        ws.insert_rows(last_row)
        ws.cell(row=last_row, column=1, value=record['Date'])
        ws.cell(row=last_row, column=2, value=record['Time In'])
        ws.cell(row=last_row, column=3, value=record['Time Out'])
        ws.cell(row=last_row, column=4, value=record['Time Spent'])
        last_row += 1
    
    # Calculate total time
    total_seconds = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1, values_only=True):
        if row[3]:  # Time Spent
            try:
                h, m, s = map(int, row[3].split(':'))  # Ensure HH:MM:SS format
                total_seconds += h * 3600 + m * 60 + s
            except ValueError:
                # Handle invalid time format (e.g., decimal values)
                pass
    
    total_time = f"{total_seconds // 3600:02}:{(total_seconds % 3600) // 60:02}:{total_seconds % 60:02}"
    
    # Update or add the total time row
    if ws.cell(row=ws.max_row, column=1).value == "Total Time:":
        ws.cell(row=ws.max_row, column=4, value=total_time)
    else:
        ws.append(['Total Time:', '', '', total_time])
    
    wb.save('time_tracker.xlsx')
    status_label.config(text="Data saved to time_tracker.xlsx!")
    update_history()

# Function to export data to Excel
def export_data():
    if not time_records:
        status_label.config(text="No data to export!")
        return
    
    # Generate a unique file name
    file_name = 'time_tracker_export.xlsx'
    counter = 1
    while True:
        try:
            with open(file_name, 'r'):
                pass
            file_name = f'time_tracker_export_{counter}.xlsx'
            counter += 1
        except FileNotFoundError:
            break
    
    wb = Workbook()
    ws = wb.active
    ws.append(['Date', 'Time In', 'Time Out', 'Time Spent'])
    total_seconds = 0
    for record in time_records:
        ws.append([record['Date'], record['Time In'], record['Time Out'], record['Time Spent']])
        if record['Time Spent']:
            try:
                h, m, s = map(int, record['Time Spent'].split(':'))  # Ensure HH:MM:SS format
                total_seconds += h * 3600 + m * 60 + s
            except ValueError:
                # Handle invalid time format (e.g., decimal values)
                pass
    total_time = f"{total_seconds // 3600:02}:{(total_seconds % 3600) // 60:02}:{total_seconds % 60:02}"
    ws.append(['Total Time:', '', '', total_time])
    wb.save(file_name)
    status_label.config(text=f"Data exported to {file_name}!")

# Function to update the history table
def update_history():
    for row in history_tree.get_children():
        history_tree.delete(row)
    for record in time_records:
        history_tree.insert("", "end", values=(record['Date'], record['Time In'], record['Time Out'], record['Time Spent']))

# Function to show history
def show_history():
    home_frame.pack_forget()
    history_frame.pack(fill=tk.BOTH, expand=True)

# Function to show home
def show_home():
    history_frame.pack_forget()
    home_frame.pack(fill=tk.BOTH, expand=True)

# Function to adjust layout based on window size
def adjust_layout(event=None):
    width = root.winfo_width()
    if width < 400:  # Small view
        label_frame.pack_forget()
        save_button.pack(side=tk.TOP, pady=10, fill=tk.X, padx=20)
        export_link.pack(side=tk.TOP, pady=5, fill=tk.X, padx=20)
        history_link.pack(side=tk.TOP, pady=5, fill=tk.X, padx=20)
    elif width < 600:  # Medium view
        label_frame.pack(fill=tk.X, pady=10)
        save_button.pack(side=tk.LEFT, pady=10, padx=20, fill=tk.X, expand=True)
        export_link.pack(side=tk.LEFT, pady=10, padx=20, fill=tk.X, expand=True)
        history_link.pack(side=tk.LEFT, pady=10, padx=20, fill=tk.X, expand=True)
    else:  # Large view
        label_frame.pack(fill=tk.X, pady=10)
        save_button.pack(side=tk.LEFT, pady=10, padx=20, fill=tk.X, expand=True)
        export_link.pack(side=tk.LEFT, pady=10, padx=20, fill=tk.X, expand=True)
        history_link.pack(side=tk.LEFT, pady=10, padx=20, fill=tk.X, expand=True)

# Create the main window
root = tk.Tk()
root.title("Time Tracker")
root.geometry("600x400")
root.configure(bg="#f0f0f0")  # Light gray background
root.minsize(300, 200)  # Minimum window size

# Style configuration
style = ttk.Style()
style.theme_use("clam")  # Use a modern theme

# Button styles
style.configure("Light.TButton", font=("Segoe UI", 12), background="#ffffff", foreground="#333333", borderwidth=0, bordercolor="#cccccc", focusthickness=3, focuscolor="#cccccc", padding=10, relief="flat", borderradius=20)
style.map("Light.TButton", background=[("active", "#f0f0f0")])

style.configure("Green.TButton", font=("Segoe UI", 12), background="#4CAF50", foreground="#ffffff", borderwidth=0, bordercolor="#4CAF50", focusthickness=3, focuscolor="#4CAF50", padding=10, relief="flat", borderradius=20)
style.map("Green.TButton", background=[("active", "#45a049")])

style.configure("Red.TButton", font=("Segoe UI", 12), background="#F44336", foreground="#ffffff", borderwidth=0, bordercolor="#F44336", focusthickness=3, focuscolor="#F44336", padding=10, relief="flat", borderradius=20)
style.map("Red.TButton", background=[("active", "#e53935")])

# Home frame
home_frame = ttk.Frame(root)
home_frame.pack(fill=tk.BOTH, expand=True)

# Top section: IN and OUT buttons
top_frame = ttk.Frame(home_frame)
top_frame.pack(fill=tk.X, pady=10)

in_button = ttk.Button(top_frame, text="IN", command=time_in, style="Light.TButton")
in_button.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10)

out_button = ttk.Button(top_frame, text="OUT", command=time_out, style="Light.TButton")
out_button.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10)

# Second section: Labels
label_frame = ttk.Frame(home_frame)
label_frame.pack(fill=tk.X, pady=10)

live_timer_label = ttk.Label(label_frame, text="Live IN Timer: 00:00:00", font=("Segoe UI", 12), background="#f0f0f0", foreground="#333333")
live_timer_label.pack(pady=5)

last_time_out_label = ttk.Label(label_frame, text="Last Time Out: None", font=("Segoe UI", 12), background="#f0f0f0", foreground="#333333")
last_time_out_label.pack(pady=5)

last_spent_time_label = ttk.Label(label_frame, text="Last Spent Time: None", font=("Segoe UI", 12), background="#f0f0f0", foreground="#333333")
last_spent_time_label.pack(pady=5)

status_label = ttk.Label(label_frame, text="Ready", font=("Segoe UI", 12), background="#f0f0f0", foreground="#333333")
status_label.pack(pady=10)

# Third section: Save, Export, and History
third_frame = ttk.Frame(home_frame)
third_frame.pack(fill=tk.X, pady=10)

save_button = ttk.Button(third_frame, text="Save", command=save_data, style="Light.TButton")
save_button.pack(side=tk.LEFT, pady=10, padx=20, fill=tk.X, expand=True)

export_link = ttk.Label(third_frame, text="Export", font=("Segoe UI", 12), foreground="blue", cursor="hand2")
export_link.pack(side=tk.LEFT, pady=10, padx=20, fill=tk.X, expand=True)
export_link.bind("<Button-1>", lambda e: export_data())

history_link = ttk.Label(third_frame, text="History", font=("Segoe UI", 12), foreground="blue", cursor="hand2")
history_link.pack(side=tk.LEFT, pady=10, padx=20, fill=tk.X, expand=True)
history_link.bind("<Button-1>", lambda e: show_history())

# History frame
history_frame = ttk.Frame(root)

history_tree = ttk.Treeview(history_frame, columns=("Date", "Time In", "Time Out", "Time Spent"), show="headings")
history_tree.heading("Date", text="Date")
history_tree.heading("Time In", text="Time In")
history_tree.heading("Time Out", text="Time Out")
history_tree.heading("Time Spent", text="Time Spent")
history_tree.pack(fill=tk.BOTH, expand=True)

back_button = ttk.Button(history_frame, text="Back", command=show_home, style="Light.TButton")
back_button.pack(side=tk.BOTTOM, pady=10, fill=tk.X, padx=20)

# Adjust layout on window resize
root.bind("<Configure>", adjust_layout)

# Initialize variables
start_time = None
timer_running = False

# Run the application
root.mainloop()