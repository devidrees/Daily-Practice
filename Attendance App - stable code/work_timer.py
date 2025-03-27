import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
from datetime import datetime, timedelta
import os
from ttkthemes import ThemedTk
import json
from PIL import Image, ImageTk
import openpyxl

class WorkTimer:
    def __init__(self):
        self.root = ThemedTk(theme="arc")
        self.root.title("Work Timer")
        self.root.geometry("1000x700")
        self.root.minsize(1000, 700)
        
        # Floating window state
        self.is_floating = False
        self.floating_window = None
        
        # Store original geometry for restore
        self.original_geometry = None
        
        # Configure grid weights for proper expansion
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)
        
        # Timer variables
        self.is_running = False
        self.start_time = None
        self.elapsed_time = timedelta()
        
        # Session history
        self.sessions = []
        
        # Settings
        self.settings = self.load_settings()
        
        # Load existing sessions if any
        self.load_sessions()
        
        # Create main container for card views
        self.main_container = ttk.Frame(self.root)
        self.main_container.grid(row=0, column=0, sticky="nsew")
        self.main_container.grid_rowconfigure(0, weight=1)
        self.main_container.grid_columnconfigure(0, weight=1)
        
        # Create frames for different views
        self.timer_view = ttk.Frame(self.main_container)
        self.settings_view = ttk.Frame(self.main_container)
        
        # Initialize sidebar components
        self.sidebar_visible = False
        self.init_sidebar()
        
        # Configure view frames for expansion
        for frame in (self.timer_view, self.settings_view):
            frame.grid_rowconfigure(0, weight=1)
            frame.grid_columnconfigure(0, weight=1)
        
        self.current_view = None
        self.create_timer_view()
        self.create_settings_view()
        self.show_timer_view()
        
        self.update_timer()
        
    def init_sidebar(self):
        # Create transparent overlay
        self.overlay = tk.Toplevel(self.root)
        self.overlay.withdraw()  # Hide initially
        self.overlay.transient(self.root)
        self.overlay.configure(bg='black')
        
        # Remove window decorations
        self.overlay.overrideredirect(True)
        self.overlay.bind('<Button-1>', self.handle_overlay_click)
        
        # Create sidebar as Toplevel
        self.sidebar = tk.Toplevel(self.root)
        self.sidebar.withdraw()  # Hide initially
        self.sidebar.transient(self.root)
        self.sidebar.overrideredirect(True)
        
        # Create main sidebar frame
        self.sidebar_frame = ttk.Frame(self.sidebar)
        self.sidebar_frame.pack(fill=tk.BOTH, expand=True)
        
        # Configure styles
        style = ttk.Style()
        style.configure('Sidebar.TFrame', background='white')
        style.configure('Sidebar.TButton', font=('Arial', 12))
        style.configure('SidebarClose.TButton', font=('Arial', 10))
        
        # Close button
        close_btn = ttk.Button(
            self.sidebar_frame,
            text="✕",
            command=self.hide_sidebar,
            style="SidebarClose.TButton",
            width=2
        )
        close_btn.pack(side=tk.TOP, anchor=tk.E, padx=5, pady=5)
        
        # Menu items container
        menu_container = ttk.Frame(self.sidebar_frame)
        menu_container.pack(fill=tk.X, padx=10, pady=10)
        
        # Settings button
        settings_btn = ttk.Button(
            menu_container,
            text="⚙ Settings",
            command=self.open_settings,
            style="Sidebar.TButton",
            width=20
        )
        settings_btn.pack(pady=5, padx=5, fill=tk.X)
        
        # About button
        about_btn = ttk.Button(
            menu_container,
            text="ℹ About",
            command=self.open_about,
            style="Sidebar.TButton",
            width=20
        )
        about_btn.pack(pady=5, padx=5, fill=tk.X)

    def open_settings(self):
        self.hide_sidebar()
        self.show_settings_view()

    def open_about(self):
        self.hide_sidebar()
        self.show_about()

    def show_sidebar(self):
        if not self.sidebar_visible:
            # Configure and show overlay
            self.overlay.deiconify()
            x = self.root.winfo_x()
            y = self.root.winfo_y()
            self.overlay.geometry(f"{self.root.winfo_width()}x{self.root.winfo_height()}+{x}+{y}")
            self.overlay.attributes('-alpha', 0.3)  # 30% opacity
            self.overlay.lift()
            
            # Show sidebar
            self.sidebar.deiconify()
            self.sidebar.geometry(f"250x{self.root.winfo_height()}+{x}+{y}")
            self.sidebar.lift()
            
            self.sidebar_visible = True

    def hide_sidebar(self):
        if self.sidebar_visible:
            # Hide components
            self.overlay.withdraw()
            self.sidebar.withdraw()
            self.sidebar_visible = False

    def handle_overlay_click(self, event):
        self.hide_sidebar()

    def show_settings_view(self):
        if self.current_view:
            self.current_view.pack_forget()
        self.settings_view.pack(fill=tk.BOTH, expand=True)
        self.current_view = self.settings_view

    def show_timer_view(self):
        if self.current_view:
            self.current_view.pack_forget()
        self.timer_view.pack(fill=tk.BOTH, expand=True)
        self.current_view = self.timer_view

    def create_timer_view(self):
        # Main frame with background color
        main_frame = ttk.Frame(self.timer_view, style='Main.TFrame')
        main_frame.grid(row=0, column=0, sticky="nsew", padx=20, pady=10)
        main_frame.grid_columnconfigure(0, weight=1)
        
        # Menu bar frame with background
        menu_frame = ttk.Frame(main_frame, style='Menu.TFrame')
        menu_frame.grid(row=0, column=0, sticky="ew", pady=(0, 20))
        menu_frame.grid_columnconfigure(1, weight=1)
        
        # Hamburger menu button - larger
        self.menu_btn = ttk.Button(
            menu_frame,
            text="☰",
            command=self.show_sidebar,
            width=3,
            style="Menu.TButton"
        )
        self.menu_btn.grid(row=0, column=0, padx=5)
        
        # Status label for notifications
        self.status_label = ttk.Label(
            menu_frame,
            text="",
            foreground="green",
            style="Status.TLabel"
        )
        self.status_label.grid(row=0, column=1, padx=10)
        
        # Float mode toggle button - larger
        self.float_btn = ttk.Button(
            menu_frame,
            text="🗗 Float",
            command=self.toggle_float_mode,
            width=8,
            style="Menu.TButton"
        )
        self.float_btn.grid(row=0, column=2, padx=5)
        
        # Timer display
        timer_frame = ttk.Frame(main_frame, style='Main.TFrame')
        timer_frame.grid(row=1, column=0, sticky="ew", pady=20)
        timer_frame.grid_columnconfigure(0, weight=1)
        
        self.timer_label = ttk.Label(
            timer_frame,
            text="00:00:00",
            font=("Arial", 64, "bold"),
            style='Timer.TLabel'
        )
        self.timer_label.grid(row=0, column=0)
        
        # Control buttons frame
        btn_frame = ttk.Frame(main_frame)
        btn_frame.grid(row=2, column=0, pady=20)
        btn_frame.grid_columnconfigure(0, weight=1)
        btn_frame.grid_columnconfigure(1, weight=1)
        
        # Control buttons with Unicode symbols - larger size
        button_configs = [
            ("▶ Start", "Green.TButton", self.start_timer),
            ("⏹ Stop", "Red.TButton", self.stop_timer)
        ]
        
        for i, (text, style, command) in enumerate(button_configs):
            btn = ttk.Button(
                btn_frame,
                text=text,
                command=command,
                style=style,
                width=20  # Increased width
            )
            btn.grid(row=0, column=i, padx=20)  # Increased padding
            
            if "Start" in text:
                self.start_btn = btn
            else:
                self.stop_btn = btn
                self.stop_btn.configure(state=tk.DISABLED)
        
        # Data management buttons
        data_frame = ttk.Frame(main_frame)
        data_frame.grid(row=3, column=0, pady=20)
        
        button_frame = ttk.Frame(data_frame)
        button_frame.pack()
        
        self.update_btn = ttk.Button(
            button_frame,
            text="⬆ Update Excel",
            command=self.update_excel,
            style="Blue.TButton",
            width=15
        )
        self.update_btn.pack(side=tk.LEFT, padx=5)
        
        # Session history
        history_frame = ttk.LabelFrame(main_frame, text="Session History", padding="10")
        history_frame.grid(row=4, column=0, sticky="nsew", pady=10)
        history_frame.grid_rowconfigure(0, weight=1)
        history_frame.grid_columnconfigure(0, weight=1)
        
        # Treeview for session history
        self.history_tree = ttk.Treeview(
            history_frame,
            columns=("Date", "Start Time", "End Time", "Duration"),
            show="headings",
            height=10
        )
        
        # Configure columns
        column_widths = {
            "Date": 150,
            "Start Time": 150,
            "End Time": 150,
            "Duration": 150
        }
        
        for col, width in column_widths.items():
            self.history_tree.heading(col, text=col)
            self.history_tree.column(col, width=width, minwidth=width)
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(history_frame, orient=tk.VERTICAL, command=self.history_tree.yview)
        self.history_tree.configure(yscrollcommand=scrollbar.set)
        
        self.history_tree.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")
        
        # Configure styles
        self.configure_styles()
        
        # Load existing sessions into the tree
        self.update_history_display()

    def create_settings_view(self):
        # Header frame with back button
        header_frame = ttk.Frame(self.settings_view)
        header_frame.grid(row=0, column=0, sticky="ew", padx=20, pady=10)
        
        back_btn = ttk.Button(
            header_frame,
            text="← Back",
            command=self.show_timer_view
        )
        back_btn.pack(side=tk.LEFT)
        
        ttk.Label(header_frame, text="Settings", font=("Arial", 20, "bold")).pack(side=tk.LEFT, padx=20)
        
        # Main settings content
        content_frame = ttk.Frame(self.settings_view, padding="20")
        content_frame.grid(row=1, column=0, sticky="nsew", padx=20)
        content_frame.grid_columnconfigure(0, weight=1)
        
        # Application Settings Section
        app_settings_frame = ttk.LabelFrame(content_frame, text="Application Settings", padding="10")
        app_settings_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Theme Selection Frame
        theme_frame = ttk.Frame(app_settings_frame)
        theme_frame.pack(fill=tk.X, pady=5)
        
        theme_label = ttk.Label(theme_frame, text="Select Theme:")
        theme_label.pack(side=tk.LEFT, padx=10)
        
        available_themes = sorted(self.root.get_themes())
        self.theme_var = tk.StringVar(value=self.settings.get('theme', 'arc'))
        
        theme_dropdown = ttk.Combobox(
            theme_frame, 
            textvariable=self.theme_var,
            values=available_themes,
            state="readonly",
            width=30
        )
        theme_dropdown.pack(side=tk.LEFT, padx=10)
        
        # Bind theme change event
        theme_dropdown.bind('<<ComboboxSelected>>', self.on_theme_change)
        
        # File Management Section
        file_settings_frame = ttk.LabelFrame(content_frame, text="File Management", padding="10")
        file_settings_frame.pack(fill=tk.X, pady=10)

        # Update File Location
        ttk.Label(file_settings_frame, text="Update Excel Location:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.update_path = ttk.Entry(file_settings_frame, width=50)
        self.update_path.grid(row=0, column=1, padx=5)
        self.update_path.insert(0, self.settings.get('update_file', 'work_log.xlsx'))
        ttk.Button(file_settings_frame, text="Browse", 
                   command=lambda: self.browse_file(self.update_path, "Excel files", ".xlsx")).grid(row=0, column=2)

        # Export File Location
        ttk.Label(file_settings_frame, text="Export File Location:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.export_path = ttk.Entry(file_settings_frame, width=50)
        self.export_path.grid(row=1, column=1, padx=5)
        self.export_path.insert(0, self.settings.get('export_file', 'work_log_export'))
        ttk.Button(file_settings_frame, text="Browse", 
                   command=lambda: self.browse_directory(self.export_path)).grid(row=1, column=2)

        # Export Format Frame
        format_frame = ttk.Frame(file_settings_frame)
        format_frame.grid(row=2, column=0, columnspan=3, pady=10, sticky="w")
        
        ttk.Label(format_frame, text="Export Format:").pack(side=tk.LEFT, padx=5)
        
        self.export_format = tk.StringVar(value=self.settings.get('export_format', 'xlsx'))
        formats = [('Excel (.xlsx)', 'xlsx'), ('CSV (.csv)', 'csv'), ('JSON (.json)', 'json')]
        
        format_options = ttk.Frame(format_frame)
        format_options.pack(side=tk.LEFT, padx=20)
        
        for i, (text, value) in enumerate(formats):
            ttk.Radiobutton(
                format_options,
                text=text,
                value=value,
                variable=self.export_format
            ).pack(side=tk.LEFT, padx=10)

        # Database Management Section
        db_frame = ttk.LabelFrame(content_frame, text="Database Management", padding="10")
        db_frame.pack(fill=tk.X, pady=10)
        
        # Warning/Alert label (will be used for both warnings and operation feedback)
        self.db_alert_label = ttk.Label(
            db_frame,
            text="⚠️ Warning: These operations can result in data loss",
            foreground="red",
            font=("Arial", 10, "bold")
        )
        self.db_alert_label.pack(pady=(0, 10))
        
        # Button container for dangerous operations
        dangerous_ops_frame = ttk.Frame(db_frame)
        dangerous_ops_frame.pack(fill=tk.X, pady=5)
        
        # Clear database button (red)
        ttk.Button(
            dangerous_ops_frame,
            text="Clear Database",
            command=self.clear_database,
            style="Danger.TButton"
        ).pack(side=tk.LEFT, padx=20)
        
        # Separator
        ttk.Separator(db_frame, orient='horizontal').pack(fill=tk.X, pady=10)
        
        # Session cleanup operations
        cleanup_frame = ttk.Frame(db_frame)
        cleanup_frame.pack(fill=tk.X, pady=5)
        
        # First row of cleanup buttons
        cleanup_row1 = ttk.Frame(cleanup_frame)
        cleanup_row1.pack(fill=tk.X, pady=2)
        
        ttk.Button(
            cleanup_row1,
            text="Remove Sessions (<5s)",
            command=self.remove_short_sessions,
            style="Database.TButton"
        ).pack(side=tk.LEFT, padx=20)
        
        ttk.Button(
            cleanup_row1,
            text="Remove Sessions (<1 min)",
            command=lambda: self.remove_sessions_under_duration(60),
            style="Database.TButton"
        ).pack(side=tk.LEFT, padx=20)
        
        # Second row of cleanup buttons
        cleanup_row2 = ttk.Frame(cleanup_frame)
        cleanup_row2.pack(fill=tk.X, pady=2)
        
        ttk.Button(
            cleanup_row2,
            text="Remove Sessions (<5 min)",
            command=lambda: self.remove_sessions_under_duration(300),
            style="Database.TButton"
        ).pack(side=tk.LEFT, padx=20)
        
        # Separator before sync section
        ttk.Separator(db_frame, orient='horizontal').pack(fill=tk.X, pady=10)
        
        # Excel Sync section
        sync_frame = ttk.Frame(db_frame)
        sync_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(
            sync_frame,
            text="Excel Synchronization:",
            font=("Arial", 10, "bold")
        ).pack(side=tk.LEFT, padx=20)
        
        ttk.Button(
            sync_frame,
            text="🔄  Sync with Database",
            command=self.sync_with_database,
            style="Sync.TButton"
        ).pack(side=tk.LEFT, padx=5)

        # Configure styles for the buttons
        style = ttk.Style()
        style.configure("Danger.TButton",
                       font=("Arial", 10, "bold"),
                       padding=(10, 5),
                       background='#dc3545',
                       foreground='white')
        
        style.configure("Database.TButton",
                       font=("Arial", 10),
                       padding=(10, 5))
        
        style.configure("Sync.TButton",
                       font=("Arial", 10, "bold"),
                       padding=(10, 5))

    def save_settings_from_view(self):
        self.settings = {
            'update_file': self.update_path.get(),
            'export_file': self.export_path.get(),
            'export_format': self.export_format.get(),
            'theme': self.theme_var.get()
        }
        with open('settings.json', 'w') as f:
            json.dump(self.settings, f)
            
        # Apply new theme
        self.root.set_theme(self.theme_var.get())
        self.show_status("Settings saved successfully")
        self.show_timer_view()

    def configure_styles(self):
        style = ttk.Style()
        
        # Configure main frame style with a light background
        style.configure('Main.TFrame', background='#f0f5ff')
        
        # Configure menu frame style
        style.configure('Menu.TFrame', background='#e6eeff')
        
        # Configure menu buttons style
        style.configure('Menu.TButton', 
                       font=('Arial', 14, 'bold'),
                       padding=(10, 5))
        
        # Configure status label style
        style.configure('Status.TLabel',
                       font=('Arial', 11),
                       background='#e6eeff')
        
        # Configure link label style
        style.configure('Link.TLabel',
                       font=('Arial', 11, 'underline'),
                       background='#e6eeff')
        
        # Configure button styles with larger font, padding and colors
        style.configure("Green.TButton", 
                       font=("Arial", 14, "bold"),
                       padding=(25, 12),
                       background='#28a745',
                       foreground='white')
        style.map("Green.TButton",
                 background=[('disabled', '#808080')],
                 foreground=[('disabled', 'white')])
        
        style.configure("Red.TButton", 
                       font=("Arial", 14, "bold"),
                       padding=(25, 12),
                       background='#dc3545',
                       foreground='white')
        style.map("Red.TButton",
                 background=[('disabled', '#808080')],
                 foreground=[('disabled', 'white')])
        
        style.configure("Blue.TButton", 
                       font=("Arial", 12),
                       padding=(15, 8),
                       background='#007bff',
                       foreground='white')
        
        # Configure floating window control buttons
        style.configure("WindowControl.TButton",
                       font=('Arial', 10, 'bold'),
                       padding=(5, 2),
                       background='#343a40',
                       foreground='white')
        
        # Configure floating window timer buttons
        style.configure("FloatingGreen.TButton",
                       font=('Arial', 12, 'bold'),
                       padding=(8, 4),
                       background='#28a745',
                       foreground='white')
        style.map("FloatingGreen.TButton",
                 background=[('disabled', '#808080')],
                 foreground=[('disabled', 'white')])
        
        style.configure("FloatingRed.TButton",
                       font=('Arial', 12, 'bold'),
                       padding=(8, 4),
                       background='#dc3545',
                       foreground='white')
        style.map("FloatingRed.TButton",
                 background=[('disabled', '#808080')],
                 foreground=[('disabled', 'white')])
        
        # Configure history frame style
        style.configure('History.TLabelframe',
                       background='#f0f5ff')
        style.configure('History.TLabelframe.Label',
                       background='#f0f5ff',
                       font=('Arial', 11, 'bold'))
        
        # Configure sidebar styles
        style.configure('Sidebar.TFrame', 
                       background='white',
                       relief='raised')
        style.configure('Sidebar.TButton',
                       font=('Arial', 12),
                       padding=(10, 5))
        style.configure('SidebarClose.TButton',
                       font=('Arial', 10),
                       padding=(5, 5))
        
        # Configure floating window styles
        style.configure('Floating.TFrame',
                       background='#f8f9fa',
                       relief='raised')
        style.configure('Floating.TButton',
                       font=('Arial', 12),
                       padding=(8, 4))
        
        # Ensure consistent layout across themes
        style.layout('TButton', [
            ('Button.padding', {'sticky': 'nswe', 'children': [
                ('Button.label', {'sticky': 'nswe'})
            ]})
        ])
        
        # Configure frame styles for consistency
        style.configure('TFrame', background=self.root.cget('background'))
        style.configure('TLabelframe', background=self.root.cget('background'))
        style.configure('TLabelframe.Label', background=self.root.cget('background'))
        
        # Add new style for database buttons
        style.configure("Database.TButton", 
                       font=("Arial", 10),
                       padding=(10, 5))
        
        # Configure timer label style to match parent background
        style.configure('Timer.TLabel',
                       font=('Arial', 64, 'bold'),
                       background='#f0f5ff')

    def start_timer(self):
        if not self.is_running:
            self.is_running = True
            self.start_time = datetime.now()
            self.elapsed_time = timedelta()
            
            # Update button states in both views
            self.start_btn.configure(state=tk.DISABLED)
            self.stop_btn.configure(state=tk.NORMAL)
            if self.is_floating and hasattr(self, 'floating_start_btn') and hasattr(self, 'floating_stop_btn'):
                self.floating_start_btn.configure(state=tk.DISABLED)
                self.floating_stop_btn.configure(state=tk.NORMAL)
    
    def stop_timer(self):
        if self.is_running:
            end_time = datetime.now()
            total_duration = end_time - self.start_time
            
            # Create and save session record
            session = {
                'date': self.start_time.strftime('%Y-%m-%d'),
                'start_time': self.start_time.strftime('%H:%M:%S'),
                'end_time': end_time.strftime('%H:%M:%S'),
                'duration': str(total_duration).split('.')[0]
            }
            
            self.sessions.append(session)
            self.save_sessions()
            self.update_history_display()
            
            # Reset all timer states
            self.is_running = False
            self.start_time = None
            self.elapsed_time = timedelta()
            
            # Reset button states in both views
            self.start_btn.configure(state=tk.NORMAL)
            self.stop_btn.configure(state=tk.DISABLED)
            if self.is_floating and hasattr(self, 'floating_start_btn') and hasattr(self, 'floating_stop_btn'):
                self.floating_start_btn.configure(state=tk.NORMAL)
                self.floating_stop_btn.configure(state=tk.DISABLED)
            
            # Reset timer display
            time_str = "00:00:00"
            self.timer_label.configure(text=time_str)
            if self.is_floating and hasattr(self, 'floating_timer_label'):
                self.floating_timer_label.configure(text=time_str)
    
    def update_timer(self):
        if self.is_running and self.start_time:
            current_time = datetime.now()
            elapsed = current_time - self.start_time
            time_str = str(elapsed).split('.')[0]
            self.timer_label.configure(text=time_str)
            if self.is_floating and self.floating_window:
                self.floating_timer_label.configure(text=time_str)
        self.root.after(1000, self.update_timer)
    
    def save_sessions(self):
        with open('sessions.json', 'w') as f:
            json.dump(self.sessions, f)
    
    def load_sessions(self):
        try:
            with open('sessions.json', 'r') as f:
                self.sessions = json.load(f)
        except FileNotFoundError:
            self.sessions = []
    
    def update_history_display(self):
        # Clear existing items
        for item in self.history_tree.get_children():
            self.history_tree.delete(item)
        
        # Add sessions to treeview
        for session in reversed(self.sessions):
            self.history_tree.insert(
                '',
                'end',
                values=(
                    session['date'],
                    session['start_time'],
                    session['end_time'],
                    session['duration']
                )
            )
    
    def update_excel(self):
        # Convert sessions to DataFrame
        df = pd.DataFrame(self.sessions)
        
        try:
            update_file = self.settings.get('update_file', 'work_log.xlsx')
            if os.path.exists(update_file):
                existing_df = pd.read_excel(update_file)
                # Remove the total row if it exists
                if not existing_df.empty and existing_df.iloc[-1]['date'].startswith('TOTAL'):
                    existing_df = existing_df.iloc[:-1]
                df = pd.concat([existing_df, df], ignore_index=True)
            
            # Convert date to datetime for sorting and day extraction
            df['date'] = pd.to_datetime(df['date'])
            
            # Add day column
            df['day'] = df['date'].dt.strftime('%A')
            
            # Sort by date
            df = df.sort_values('date')
            
            # Remove entries with zero duration
            df = df[df['duration'] != '0:00:00']
            
            # Format date back to string
            df['date'] = df['date'].dt.strftime('%Y-%m-%d')
            
            # Reorder columns to put day after date
            df = df[['date', 'day', 'start_time', 'end_time', 'duration']]
            
            # Group by month for totals
            df['month'] = pd.to_datetime(df['date']).dt.strftime('%Y-%m')
            monthly_groups = df.groupby('month')
            
            # Create final DataFrame with monthly totals
            final_dfs = []
            for month, group in monthly_groups:
                # Calculate monthly total duration in seconds
                total_seconds = sum(
                    pd.Timedelta(duration).total_seconds() 
                    for duration in group['duration']
                )
                
                # Convert to hours and minutes
                total_hours = int(total_seconds // 3600)
                remaining_minutes = int((total_seconds % 3600) // 60)
                
                # Format the duration string
                if total_hours == 0:
                    duration_str = f"{remaining_minutes} minutes"
                elif total_hours == 1:
                    if remaining_minutes == 0:
                        duration_str = "1 hour"
                    else:
                        duration_str = f"1 hour {remaining_minutes} minutes"
                else:
                    if remaining_minutes == 0:
                        duration_str = f"{total_hours} hours"
                    else:
                        duration_str = f"{total_hours} hours {remaining_minutes} minutes"
                
                # Add the group's data
                group_df = group.drop('month', axis=1)
                
                # Add monthly total row
                month_name = pd.to_datetime(month + '-01').strftime('%B %Y')
                total_style = pd.DataFrame([{
                    'date': f'TOTAL FOR {month_name}',
                    'day': '',
                    'start_time': '',
                    'end_time': '',
                    'duration': duration_str
                }])
                
                # Combine group data with its total
                final_dfs.append(pd.concat([group_df, total_style], ignore_index=True))
            
            # Combine all monthly groups
            df = pd.concat(final_dfs, ignore_index=True)
            
            # Create Excel writer with formatting
            with pd.ExcelWriter(update_file, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Sheet1']
                
                # Format all monthly total rows
                for row in range(1, worksheet.max_row + 1):
                    if str(worksheet.cell(row=row, column=1).value).startswith('TOTAL FOR'):
                        # Apply formatting to the total row
                        for col in range(1, 6):  # A to E columns (including day column)
                            cell = worksheet.cell(row=row, column=col)
                            cell.font = openpyxl.styles.Font(bold=True, size=12)
                            cell.fill = openpyxl.styles.PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
                            if col == 1:  # 'TOTAL FOR' cell
                                cell.alignment = openpyxl.styles.Alignment(horizontal='left')
                            elif col == 5:  # Duration cell
                                cell.alignment = openpyxl.styles.Alignment(horizontal='right')
                        
                        # Add bottom border to the row above total
                        for col in range(1, 6):
                            cell = worksheet.cell(row=row-1, column=col)
                            cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))
                
                # Adjust column widths
                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column].width = adjusted_width
            
            self.show_status("Work log updated successfully!", 5000, update_file)
        except Exception as e:
            self.show_status(f"Error: {str(e)}", 5000)
    
    def browse_file(self, entry_widget, file_type, extension):
        filename = filedialog.asksaveasfilename(
            defaultextension=extension,
            filetypes=[(file_type, f"*{extension}")]
        )
        if filename:
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, filename)

    def browse_directory(self, entry_widget):
        directory = filedialog.askdirectory()
        if directory:
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, directory)

    def load_settings(self):
        try:
            with open('settings.json', 'r') as f:
                settings = json.load(f)
                # Apply theme from settings
                if 'theme' in settings:
                    self.root.set_theme(settings['theme'])
                return settings
        except FileNotFoundError:
            return {
                'update_file': 'work_log.xlsx',
                'export_file': 'work_log_export',
                'export_format': 'xlsx',
                'theme': 'arc'
            }

    def show_status(self, message, duration=3000, file_path=None):
        if file_path:
            # Create a frame for status and link
            status_frame = ttk.Frame(self.menu_btn.master)
            status_frame.grid(row=0, column=1, padx=10)
            
            # Status message
            msg_label = ttk.Label(
                status_frame,
                text=message,
                foreground="green",
                style="Status.TLabel"
            )
            msg_label.pack(side=tk.LEFT)
            
            # Link to file
            link_label = ttk.Label(
                status_frame,
                text="Open File",
                foreground="blue",
                cursor="hand2",
                style="Link.TLabel"
            )
            link_label.pack(side=tk.LEFT, padx=(5, 0))
            
            # Bind click event
            link_label.bind("<Button-1>", lambda e: self.open_file(file_path))
            
            # Store reference to prevent garbage collection
            self.status_label = status_frame
            
            # Clear after duration
            self.root.after(duration, lambda: self.clear_status())
        else:
            # Regular status message
            self.status_label.configure(text=message)
            self.root.after(duration, lambda: self.status_label.configure(text=""))

    def clear_status(self):
        if isinstance(self.status_label, ttk.Frame):
            self.status_label.destroy()
            # Recreate the original status label
            self.status_label = ttk.Label(
                self.menu_btn.master,
                text="",
                foreground="green",
                style="Status.TLabel"
            )
            self.status_label.grid(row=0, column=1, padx=10)

    def open_file(self, file_path):
        try:
            os.startfile(file_path) if os.name == 'nt' else os.system(f'open "{file_path}"')
        except Exception as e:
            self.show_status(f"Error opening file: {str(e)}", 5000)

    def show_about(self):
        messagebox.showinfo("About", "Work Timer v1.0\nA simple time tracking application.")

    def on_theme_change(self, event=None):
        # Store current geometry
        current_geometry = self.root.geometry()
        
        # Apply new theme
        self.root.set_theme(self.theme_var.get())
        
        # Reconfigure styles for consistency
        self.configure_styles()
        
        # Update timer label background to match new theme
        bg_color = style.lookup('Main.TFrame', 'background')
        style.configure('Timer.TLabel', background=bg_color)
        
        # Restore geometry
        self.root.geometry(current_geometry)
        
        # Update status
        self.show_status("Theme updated")

    def create_floating_window(self):
        self.floating_window = tk.Toplevel(self.root)
        self.floating_window.overrideredirect(True)
        self.floating_window.attributes('-topmost', True)
        
        # Create main frame with border effect
        main_frame = ttk.Frame(self.floating_window, style='Floating.TFrame')
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Title bar with dark background
        title_bar = ttk.Frame(main_frame, style='Title.TFrame')
        title_bar.pack(fill=tk.X, padx=2, pady=2)
        title_bar.grid_columnconfigure(0, weight=1)
        
        # Bind dragging events
        title_bar.bind('<Button-1>', self.start_move)
        title_bar.bind('<B1-Motion>', self.on_move)
        
        # Window title
        ttk.Label(
            title_bar,
            text="Work Timer",
            font=('Arial', 10, 'bold'),
            style='Title.TLabel'
        ).grid(row=0, column=0, padx=5)
        
        # Window controls frame
        controls_frame = ttk.Frame(title_bar, style='Title.TFrame')
        controls_frame.grid(row=0, column=1, padx=2)
        
        # Window control buttons with proper styling
        ttk.Button(
            controls_frame,
            text="−",
            width=2,
            command=self.floating_window.iconify,
            style="WindowControl.TButton"
        ).pack(side=tk.LEFT, padx=1)
        
        ttk.Button(
            controls_frame,
            text="⬚",
            width=2,
            command=self.toggle_float_mode,
            style="WindowControl.TButton"
        ).pack(side=tk.LEFT, padx=1)
        
        ttk.Button(
            controls_frame,
            text="✕",
            width=2,
            command=self.toggle_float_mode,
            style="WindowControl.TButton"
        ).pack(side=tk.LEFT, padx=1)
        
        # Timer content
        content_frame = ttk.Frame(main_frame, padding=10)
        content_frame.pack(fill=tk.BOTH, expand=True)
        
        # Timer display
        self.floating_timer_label = ttk.Label(
            content_frame,
            text="00:00:00",
            font=("Arial", 24, "bold")  # Increased font size
        )
        self.floating_timer_label.pack(pady=5)
        
        # Control buttons
        btn_frame = ttk.Frame(content_frame)
        btn_frame.pack(pady=5)
        
        # Start button with green style
        self.floating_start_btn = ttk.Button(
            btn_frame,
            text="▶",
            command=self.start_timer,
            width=3,
            style="FloatingGreen.TButton"
        )
        self.floating_start_btn.pack(side=tk.LEFT, padx=2)
        
        # Stop button with red style
        self.floating_stop_btn = ttk.Button(
            btn_frame,
            text="⏹",
            command=self.stop_timer,
            width=3,
            style="FloatingRed.TButton",
            state=tk.DISABLED
        )
        self.floating_stop_btn.pack(side=tk.LEFT, padx=2)
        
        # Configure title bar styles
        style = ttk.Style()
        style.configure('Title.TFrame', background='#343a40')
        style.configure('Title.TLabel', foreground='white', background='#343a40')
        
        # Set initial position
        self.floating_window.geometry('220x140+100+100')  # Slightly larger for better visibility

    def start_move(self, event):
        self.x = event.x
        self.y = event.y

    def on_move(self, event):
        deltax = event.x - self.x
        deltay = event.y - self.y
        x = self.floating_window.winfo_x() + deltax
        y = self.floating_window.winfo_y() + deltay
        self.floating_window.geometry(f"+{x}+{y}")

    def toggle_float_mode(self):
        if not self.is_floating:
            # Store current geometry and switch to floating mode
            self.original_geometry = self.root.geometry()
            self.root.withdraw()
            self.create_floating_window()
            self.is_floating = True
        else:
            # Restore main window
            if self.floating_window:
                self.floating_window.destroy()
                self.floating_window = None
            self.root.deiconify()
            if self.original_geometry:
                self.root.geometry(self.original_geometry)
            self.is_floating = False

    def run(self):
        self.root.mainloop()

    def show_db_alert(self, message, is_warning=False):
        self.db_alert_label.configure(
            text=message,
            foreground="red" if is_warning else "green"
        )
        if not is_warning:
            # Reset to warning message after 5 seconds
            self.root.after(5000, lambda: self.db_alert_label.configure(
                text="⚠️ Warning: These operations can result in data loss",
                foreground="red"
            ))

    def clear_database(self):
        if messagebox.askyesno("Confirm Clear", "Are you sure you want to clear all session data? This cannot be undone."):
            self.sessions = []
            self.save_sessions()
            self.update_history_display()
            self.show_db_alert("Database cleared successfully!")

    def remove_short_sessions(self):
        original_count = len(self.sessions)
        self.sessions = [
            session for session in self.sessions 
            if pd.Timedelta(session['duration']).total_seconds() >= 5
        ]
        removed_count = original_count - len(self.sessions)
        self.save_sessions()
        self.update_history_display()
        self.show_db_alert(f"Removed {removed_count} sessions under 5 seconds")

    def remove_sessions_under_duration(self, seconds):
        original_count = len(self.sessions)
        self.sessions = [
            session for session in self.sessions 
            if pd.Timedelta(session['duration']).total_seconds() >= seconds
        ]
        removed_count = original_count - len(self.sessions)
        self.save_sessions()
        self.update_history_display()
        
        duration_str = f"{seconds//60} minute{'s' if seconds >= 120 else ''}"
        self.show_db_alert(f"Removed {removed_count} sessions under {duration_str}")

    def sync_with_database(self):
        try:
            update_file = self.settings.get('update_file', 'work_log.xlsx')
            if not os.path.exists(update_file):
                self.show_db_alert("Excel file not found. Please update Excel first.", is_warning=True)
                return
                
            # Read existing Excel file
            existing_df = pd.read_excel(update_file)
            
            # Store the total rows
            total_rows = existing_df[existing_df['date'].str.startswith('TOTAL FOR', na=False)]
            
            # Remove total rows from existing data
            existing_df = existing_df[~existing_df['date'].str.startswith('TOTAL FOR', na=False)]
            
            # Create DataFrame from current database
            db_df = pd.DataFrame(self.sessions)
            if not db_df.empty:
                # Convert date to datetime for comparison
                db_df['date'] = pd.to_datetime(db_df['date'])
                existing_df['date'] = pd.to_datetime(existing_df['date'])
                
                # Create a unique identifier for each record (date + start_time + end_time)
                existing_df['record_id'] = existing_df['date'].dt.strftime('%Y-%m-%d') + '_' + existing_df['start_time'] + '_' + existing_df['end_time']
                db_records = set(
                    pd.to_datetime(s['date']).strftime('%Y-%m-%d') + '_' + 
                    s['start_time'] + '_' + 
                    s['end_time'] 
                    for s in self.sessions
                )
                
                # Keep only records that exist in database
                synced_df = existing_df[existing_df['record_id'].isin(db_records)]
                removed_count = len(existing_df) - len(synced_df)
                
                if not synced_df.empty:
                    # Drop the temporary record_id column
                    synced_df = synced_df.drop('record_id', axis=1)
                    
                    # Convert date back to string
                    synced_df['date'] = synced_df['date'].dt.strftime('%Y-%m-%d')
                    
                    # Add day column if not present
                    if 'day' not in synced_df.columns:
                        synced_df['date'] = pd.to_datetime(synced_df['date'])
                        synced_df['day'] = synced_df['date'].dt.strftime('%A')
                        synced_df['date'] = synced_df['date'].dt.strftime('%Y-%m-%d')
                    
                    # Reorder columns
                    synced_df = synced_df[['date', 'day', 'start_time', 'end_time', 'duration']]
                    
                    # Combine with total rows
                    final_df = pd.concat([synced_df, total_rows], ignore_index=True)
                    
                    # Save back to Excel with formatting
                    with pd.ExcelWriter(update_file, engine='openpyxl') as writer:
                        final_df.to_excel(writer, index=False)
                        
                        # Get the workbook and worksheet
                        workbook = writer.book
                        worksheet = writer.sheets['Sheet1']
                        
                        # Format total rows
                        for row in range(1, worksheet.max_row + 1):
                            if str(worksheet.cell(row=row, column=1).value).startswith('TOTAL FOR'):
                                for col in range(1, 6):
                                    cell = worksheet.cell(row=row, column=col)
                                    cell.font = openpyxl.styles.Font(bold=True, size=12)
                                    cell.fill = openpyxl.styles.PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
                                    if col == 1:
                                        cell.alignment = openpyxl.styles.Alignment(horizontal='left')
                                    elif col == 5:
                                        cell.alignment = openpyxl.styles.Alignment(horizontal='right')
                        
                        # Adjust column widths
                        for col in worksheet.columns:
                            max_length = 0
                            column = col[0].column_letter
                            for cell in col:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                                adjusted_width = (max_length + 2)
                                worksheet.column_dimensions[column].width = adjusted_width
                    
                    status_msg = f"Excel synced with database. Removed {removed_count} records not in database."
                    self.show_db_alert(status_msg)
                else:
                    self.show_db_alert("No matching records found in database")
            else:
                # If database is empty, create empty Excel with just headers
                empty_df = pd.DataFrame(columns=['date', 'day', 'start_time', 'end_time', 'duration'])
                with pd.ExcelWriter(update_file, engine='openpyxl') as writer:
                    empty_df.to_excel(writer, index=False)
                self.show_db_alert("Database is empty. Created empty Excel file.")
        except Exception as e:
            self.show_db_alert(f"Error during sync: {str(e)}", is_warning=True)

if __name__ == "__main__":
    app = WorkTimer()
    app.run() 